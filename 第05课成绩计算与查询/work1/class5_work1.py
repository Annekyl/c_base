import openpyxl


def write(file, sheet):
    with open(file, "w") as f:
        for i in range(1, sheet.max_row):
            for j in range(1, 9):
                if sheet.cell(i, j).value:
                    f.write(str(sheet.cell(i, j).value))
                    f.write("\t")
                else:
                    f.write("0")
                    f.write("\t")
            f.write("\n")


# 打开XLSX文件
system = openpyxl.load_workbook('（前3次）2022级导论考试--成绩表.xlsx')

# 选择第一个工作表
sheet1 = system["598854"]
sheet2 = system["598856"]
sheet3 = system["598858"]
write("598854.txt", sheet1)
write("598856.txt", sheet2)
write("598858.txt", sheet3)

# 关闭XLSX文件
system.close()
